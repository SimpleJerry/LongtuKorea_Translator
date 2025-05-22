# Imports the Google Cloud Translation library
import math
import os
from abc import ABCMeta, abstractmethod

import pandas as pd
from PyQt5.QtCore import pyqtSignal
from openpyxl import load_workbook
from transformers import AutoModelForSeq2SeqLM, AutoTokenizer, pipeline


def translate_texts(
        texts: list,
        translator: pipeline,
        progress_bar_init: pyqtSignal(int) = None,
        progress_bar_num: pyqtSignal(int) = None,
) -> list:
    """
    :param texts: your text list to translate
    :param translator: transformers.pipeline
    :param progress_bar_init: signal to init the progressbar
    :param progress_bar_num: signal to update the value of progressbar
    :return:
    """

    len_texts = len(texts)
    # 进度条
    if progress_bar_init is not None:
        progress_bar_init.emit(len_texts)
    # 翻译
    texts_translated = []
    for index, text in enumerate(texts):
        if progress_bar_num is not None:
            progress_bar_num.emit(index)
        text_translated = translator(text)[0]['translation_text']
        texts_translated.append(text_translated)

    # TODO:
    # df = pd.read_csv(glossary_filepath)
    # glossary_dict = {index: {"zh-CN": row["zh-CN"], "ko": row["ko"]} for index, row in df.iterrows()}

    # # pre-process
    # texts_tagged = []
    # for text in texts:
    #     for glossary_id, glossary in glossary_dict.items():
    #         text = text.replace(glossary["zh-CN"], "<glossary_id={0}>".format(glossary_id))
    #     texts_tagged.append(text)
    # # translate4
    # translated = translator(texts_tagged)
    # texts_tagged_translated = [item['translation_text'] for item in translated]
    # # post-process
    # texts_translated = []
    # for text_tagged in texts_tagged_translated:
    #     glossary_ids = re.findall(r"<glossary_id=(\d+)>", text_tagged)
    #     glossary_ids = [int(glossary_id) for glossary_id in glossary_ids]
    #     for glossary_id in glossary_ids:
    #         text_tagged = text_tagged.replace("<glossary_id={0}>".format(glossary_id), glossary_dict[glossary_id]["ko"])
    #     texts_translated.append(text_tagged)

    return texts_translated


class FileTranslatorInterface(metaclass=ABCMeta):
    @abstractmethod
    def translate(self,
                  file_path_source: str,
                  translator: pipeline,
                  progress_bar_init: pyqtSignal(int) = None,
                  progress_bar_num: pyqtSignal(int) = None) -> None:
        """
        Translate txt file with glossary and progress bar.
        :param file_path_source: source file path
        :param translator: transformers.pipeline
        :param progress_bar_init: signal to init the progressbar
        :param progress_bar_num: signal to update the value of progressbar
        :return:
        """
        pass


class ExcelFileTranslator(FileTranslatorInterface):

    @staticmethod
    def _get_translatable_texts(texts: list[str]):
        indexes = []
        texts_to_translate = []
        for index, text in enumerate(texts):
            if isinstance(text, str):
                indexes.append(index)
                texts_to_translate.append(text)
        return indexes, texts_to_translate

    @staticmethod
    def _map_translated_texts(indexes, texts_translated: list[str], texts: list[str]):
        texts_translated_dict = dict(zip(indexes, texts_translated))
        for index in indexes:
            texts[index] = texts_translated_dict[index]
        return texts

    def translate(self,
                  file_path_source: str,
                  translator: pipeline,
                  progress_bar_init: pyqtSignal(int) = None,
                  progress_bar_num: pyqtSignal(int) = None) -> None:

        # Load the source workbook
        workbook = load_workbook(filename=file_path_source)

        # 过滤并翻译
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            cells = []
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cells.append(cell)

            texts = [str(cell.value) for cell in cells]
            # 处理
            indexes, texts_to_translate = self._get_translatable_texts(texts)
            texts_translated = translate_texts(texts_to_translate,
                                               translator,
                                               progress_bar_init=progress_bar_init,
                                               progress_bar_num=progress_bar_num)
            # reshape
            texts = self._map_translated_texts(indexes, texts_translated, texts)

            for i in range(len(cells)):
                cells[i].value = texts[i]

        # Save the destination workbook
        file_name, file_type = os.path.splitext(file_path_source)
        file_path_target = "{0}_translated{1}".format(file_name, file_type)
        workbook.save(filename=file_path_target)


class BaseFileTranslator(FileTranslatorInterface):
    def load_file(self, file_path_source: str):
        raise NotImplementedError

    def save_file(self, df, file_path_source: str):
        raise NotImplementedError

    def translate(self,
                  file_path_source: str,
                  translator: pipeline,
                  progress_bar_init: pyqtSignal(int) = None,
                  progress_bar_num: pyqtSignal(int) = None) -> None:
        # Load the source workbook
        df = self.load_file(file_path_source)

        cells = []
        for row_index, row in df.iterrows():
            for column_index in df.columns:
                cell_value = row[column_index]
                if isinstance(cell_value, float) and math.isnan(cell_value):
                    continue
                cells.append([cell_value, (row_index, column_index)])

        # Translate
        texts = [str(item[0]) for item in cells]
        coordinates = [item[1] for item in cells]
        texts_translated = translate_texts(texts, translator,
                                           progress_bar_init=progress_bar_init,
                                           progress_bar_num=progress_bar_num)
        # working
        for index in range(len(texts)):
            x, y = coordinates[index][0], coordinates[index][1]
            df.at[x, y] = texts_translated[index]

        # Save the destination workbook
        self.save_file(df, file_path_source)


class CsvFileTranslator(BaseFileTranslator):
    def load_file(self, file_path_source: str):
        return pd.read_csv(file_path_source)

    def save_file(self, df, file_path_source: str):
        file_name, file_type = os.path.splitext(file_path_source)
        file_path_target = "{0}_translated{1}".format(file_name, file_type)
        df.to_csv(file_path_target, index=False)


class TsvFileTranslator(BaseFileTranslator):
    def load_file(self, file_path_source: str):
        return pd.read_csv(file_path_source, sep="\t")

    def save_file(self, df, file_path_source: str):
        file_name, file_type = os.path.splitext(file_path_source)
        file_path_target = "{0}_translated{1}".format(file_name, file_type)
        df.to_csv(file_path_target, index=False, sep='\t')


class TxtFileTranslator(FileTranslatorInterface):
    def translate(self,
                  file_path_source: str,
                  translator: pipeline,
                  progress_bar_init: pyqtSignal(int) = None,
                  progress_bar_num: pyqtSignal(int) = None) -> None:
        # read
        with open(file_path_source, "r", encoding="utf-8") as file_input:
            lines = file_input.readlines()
            texts = [line for line in lines]
        # translate
        translated_texts = translate_texts(texts, translator,
                                           progress_bar_init=progress_bar_init,
                                           progress_bar_num=progress_bar_num)

        # write
        file_name, file_type = os.path.splitext(file_path_source)
        file_path_target = "{0}_translated{1}".format(file_name, file_type)
        with open(file_path_target, "w", encoding="utf-8") as file_output:
            file_output.writelines([text.rstrip("\n") + "\n" for text in translated_texts])


# Simple Factory Pattern
class FileTranslateFactory:
    @staticmethod
    def create_document_translator(file_type: str) -> FileTranslatorInterface:
        # Excel File
        if file_type == "xlsx":
            return ExcelFileTranslator()
        # csv File
        elif file_type == "csv":
            return CsvFileTranslator()
        # tsv File
        elif file_type == "tsv":
            return TsvFileTranslator()
        # txt File
        elif file_type == "txt":
            return TxtFileTranslator()
        else:
            raise TypeError("暂未支持的文件类型:{0}".format(file_type))


if __name__ == '__main__':
    model_name = r"D:\LongtuKoreaTranslationModel\fine-tuned-models\nllb-200-1.3B\zh2ko_0907"
    model = AutoModelForSeq2SeqLM.from_pretrained(model_name)
    tokenizer = AutoTokenizer.from_pretrained(model_name)
    translator = pipeline("translation", model=model, tokenizer=tokenizer,
                          src_lang=tokenizer.src_lang, tgt_lang=tokenizer.tgt_lang,
                          device="cuda:0")

    results = translate_texts(["你好", "世界"], translator)
    pass
