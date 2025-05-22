# Imports the Google Cloud Translation library
import html
import logging
import math
import os
from abc import ABCMeta, abstractmethod

import pandas as pd
from PyQt5.QtCore import pyqtSignal
from google.cloud import translate_v3 as translate
from openpyxl import load_workbook

from settings import *


def translate_texts(
        texts: list,
        project_id: str,
        source_language_code: str,
        target_language_code: str,
        glossary_id: str = None,
        progress_bar_init: pyqtSignal(int) = None,
        progress_bar_num: pyqtSignal(int) = None,
        batch_size: int = 200,
        location: str = "us-central1",
) -> list:
    """
    Translate text with glossary.
    :param texts: your text list to translate
    :param project_id: your project id
    :param source_language_code: source language code
    :param target_language_code: target language code
    :param glossary_id: your glossary id
    :param progress_bar_init: signal to init the progressbar
    :param progress_bar_num: signal to update the value of progressbar
    :param batch_size: the size of every batch of the texts to be translated
    :param location: the location of Google Translation API Resources(you don't need to modify it)
    :return:
    """

    client = translate.TranslationServiceClient()
    parent = f"projects/{project_id}/locations/{location}"

    # 用语集参数
    glossary_config = None
    if glossary_id is not None:
        glossary = client.glossary_path(
            project_id, "us-central1", glossary_id  # The location of the glossary
        )
        glossary_config = translate.TranslateTextGlossaryConfig(glossary=glossary)

    # 接口调用
    texts_translated = []
    len_texts = len(texts)
    # 进度条
    if progress_bar_init is not None:
        progress_bar_init.emit(len_texts)
    for index in range(0, len_texts, batch_size):
        if progress_bar_num is not None:
            progress_bar_num.emit(index)
        batch_texts = texts[index:min(index + batch_size, len_texts)]
        response = client.translate_text(
            request={
                "contents": batch_texts,
                "source_language_code": source_language_code,
                "target_language_code": target_language_code,
                "parent": parent,
                "glossary_config": glossary_config,
            }
        )
        if glossary_id is not None:
            for response_translation in response.glossary_translations:
                texts_translated.append(html.unescape(response_translation.translated_text))
        else:
            for response_translation in response.translations:
                texts_translated.append(html.unescape(response_translation.translated_text))
        logging.info("Translated text: {0}".format(batch_texts))

    return texts_translated


class FileTranslatorInterface(metaclass=ABCMeta):
    @abstractmethod
    def translate(self,
                  file_path_source: str,
                  project_id: str,
                  source_language_code: str,
                  target_language_code: str,
                  glossary_id: str = None,
                  progress_bar_init: pyqtSignal(int) = None,
                  progress_bar_num: pyqtSignal(int) = None) -> None:
        """
        Translate txt file with glossary and progress bar.
        :param file_path_source: source file path
        :param project_id: your project id
        :param source_language_code: source language code
        :param target_language_code: target language code
        :param glossary_id: your glossary id
        :param progress_bar_init: signal to init the progressbar
        :param progress_bar_num: signal to update the value of progressbar
        :return:
        """
        pass


class ExcelFileTranslator(FileTranslatorInterface):

    @staticmethod
    def _get_translatable_texts(texts: list[str]):
        indexes = []
        texts_to_translate = []
        for index, text in enumerate(texts):
            if isinstance(text, str):
                indexes.append(index)
                texts_to_translate.append(text)
        return indexes, texts_to_translate

    @staticmethod
    def _map_translated_texts(indexes, texts_translated: list[str], texts: list[str]):
        texts_translated_dict = dict(zip(indexes, texts_translated))
        for index in indexes:
            texts[index] = texts_translated_dict[index]
        return texts

    def translate(self,
                  file_path_source: str,
                  project_id: str,
                  source_language_code: str,
                  target_language_code: str,
                  glossary_id: str = None,
                  progress_bar_init: pyqtSignal(int) = None,
                  progress_bar_num: pyqtSignal(int) = None) -> None:

        # Load the source workbook
        workbook = load_workbook(filename=file_path_source)

        # 过滤并翻译
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            cells = []
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cells.append(cell)

            texts = [str(cell.value) for cell in cells]
            # 处理
            indexes, texts_to_translate = self._get_translatable_texts(texts)
            texts_translated = translate_texts(texts_to_translate, project_id,
                                               source_language_code,
                                               target_language_code,
                                               glossary_id,
                                               progress_bar_init=progress_bar_init,
                                               progress_bar_num=progress_bar_num)
            # reshape
            texts = self._map_translated_texts(indexes, texts_translated, texts)

            for i in range(len(cells)):
                cells[i].value = texts[i]

        # Save the destination workbook
        file_name, file_type = os.path.splitext(file_path_source)
        file_path_target = "{0}_translated{1}".format(file_name, file_type)
        workbook.save(filename=file_path_target)


class BaseFileTranslator(FileTranslatorInterface):
    def load_file(self, file_path_source: str):
        raise NotImplementedError

    def save_file(self, df, file_path_source: str):
        raise NotImplementedError

    def translate(self,
                  file_path_source: str,
                  project_id: str,
                  source_language_code: str,
                  target_language_code: str,
                  glossary_id: str = None,
                  progress_bar_init: pyqtSignal(int) = None,
                  progress_bar_num: pyqtSignal(int) = None) -> None:
        # Load the source workbook
        df = self.load_file(file_path_source)

        cells = []
        for row_index, row in df.iterrows():
            for column_index in df.columns:
                cell_value = row[column_index]
                if isinstance(cell_value, float) and math.isnan(cell_value):
                    continue
                cells.append([cell_value, (row_index, column_index)])

        # Translate
        texts = [str(item[0]) for item in cells]
        coordinates = [item[1] for item in cells]
        texts_translated = translate_texts(texts, project_id,
                                           source_language_code,
                                           target_language_code,
                                           glossary_id,
                                           progress_bar_init=progress_bar_init,
                                           progress_bar_num=progress_bar_num)
        # working
        for index in range(len(texts)):
            x, y = coordinates[index][0], coordinates[index][1]
            df.at[x, y] = texts_translated[index]

        # Save the destination workbook
        self.save_file(df, file_path_source)


class CsvFileTranslator(BaseFileTranslator):
    def load_file(self, file_path_source: str):
        return pd.read_csv(file_path_source)

    def save_file(self, df, file_path_source: str):
        file_name, file_type = os.path.splitext(file_path_source)
        file_path_target = "{0}_translated{1}".format(file_name, file_type)
        df.to_csv(file_path_target, index=False)


class TsvFileTranslator(BaseFileTranslator):
    def load_file(self, file_path_source: str):
        return pd.read_csv(file_path_source, sep="\t")

    def save_file(self, df, file_path_source: str):
        file_name, file_type = os.path.splitext(file_path_source)
        file_path_target = "{0}_translated{1}".format(file_name, file_type)
        df.to_csv(file_path_target, index=False, sep='\t')


class TxtFileTranslator(FileTranslatorInterface):
    def translate(self,
                  file_path_source: str,
                  project_id: str,
                  source_language_code: str,
                  target_language_code: str,
                  glossary_id: str = None,
                  progress_bar_init: pyqtSignal(int) = None,
                  progress_bar_num: pyqtSignal(int) = None) -> None:
        # read
        with open(file_path_source, "r", encoding="utf-8") as file_input:
            lines = file_input.readlines()
            texts = [line for line in lines]
        # translate
        translated_texts = translate_texts(texts, project_id,
                                           source_language_code,
                                           target_language_code,
                                           glossary_id,
                                           progress_bar_init=progress_bar_init,
                                           progress_bar_num=progress_bar_num)

        # write
        file_name, file_type = os.path.splitext(file_path_source)
        file_path_target = "{0}_translated{1}".format(file_name, file_type)
        with open(file_path_target, "w", encoding="utf-8") as file_output:
            file_output.writelines(translated_texts)


# Simple Factory Pattern
class FileTranslateFactory:
    @staticmethod
    def create_document_translator(file_type: str) -> FileTranslatorInterface:
        # Excel File
        if file_type == "xlsx":
            return ExcelFileTranslator()
        # csv File
        elif file_type == "csv":
            return CsvFileTranslator()
        # tsv File
        elif file_type == "tsv":
            return TsvFileTranslator()
        # txt File
        elif file_type == "txt":
            return TxtFileTranslator()
        else:
            raise TypeError("暂未支持的文件类型:{0}".format(file_type))


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    # 设置环境变量
    os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = os.path.abspath("./front/{0}".format(PRIVATE_KEY_NAME))
    excel_translator = ExcelFileTranslator()
    csv_translator = CsvFileTranslator()
    txt_translator = TxtFileTranslator()
    # text_test = ["你好"]
    # translate_texts(text_test, PROJECT_ID, "zh-CN", "ko")  # 翻译文本_带术语表（高级版）
    # translate_texts(text_test, PROJECT_ID, "zh-CN", "ko", GLOSSARY_ID_ALL)  # 翻译文本_带术语表（高级版）
    # excel_translator.translate("test/test.xlsx", PROJECT_ID, "zh-CN", "ko", GLOSSARY_ID_ALL)
    # csv_translator.translate("test/test.csv", PROJECT_ID, "zh-CN", "ko", GLOSSARY_ID_ALL)
    # txt_translator.translate("test/test.txt", PROJECT_ID, "zh-CN", "ko", GLOSSARY_ID_ALL)
    # txt_translator.translate("test/test.txt", PROJECT_ID, "zh-CN", "ko", GLOSSARY_ID_ALL)
